import openpyxl
import os

def procesar_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    resumen = wb["Resumen"]

    # Simulamos edición en columnas C, D, L, M (índices 3, 4, 12, 13)
    for i in range(2, resumen.max_row + 1):
        resumen.cell(row=i, column=3).value = resumen.cell(row=i, column=3).value or "EDITADO"
        resumen.cell(row=i, column=4).value = resumen.cell(row=i, column=4).value or "EDITADO"
        resumen.cell(row=i, column=12).value = resumen.cell(row=i, column=12).value or "OK"
        resumen.cell(row=i, column=13).value = resumen.cell(row=i, column=13).value or "OK"

    # Guardamos el nuevo archivo generado
    output_path = os.path.join("uploads", "formulario_generado.xlsx")
    wb.save(output_path)
    return output_path
